import pandas as pd
import csv
from SetUpFile import *
from fuzzywuzzy import fuzz
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import numpy as np
import numbers

#region excel load

def csv_until_emptyline(filepath):

    with open(filepath, newline = '') as f:
        contents = []
        r = csv.reader(f)

        for i, l in enumerate(r):

            if not l:
                break

            if i == 0:
                headers = l
                continue

            contents.append(l)


    return pd.DataFrame(data = contents, columns = headers)

def xlsx_until_emptyline(filepath):
    df = pd.read_excel(filepath)

    try:
        first_row_with_all_NaN = df[df.isnull().all(axis = 1) == True].index.tolist()[0]

        return df.loc[0: first_row_with_all_NaN - 1]

    except IndexError as er:

        return df

def from_dct_to_xlsx(res_dct, res_filepath):
    res_book = openpyxl.Workbook()
    res_sheet = res_book.active

    start_row = 1

    temp_row = None

    for key in res_dct:

        res_sheet.cell(row=start_row, column=1, value=key)
        start_row += 1

        rows = list(dataframe_to_rows(res_dct[key]))
        rows.pop(1)

        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                res_sheet.cell(row=start_row + r_idx, column=c_idx, value=value)
            temp_row = r_idx

        start_row = temp_row + start_row + 2

    res_book.save(res_filepath)
    res_book.close()

def process_column_type(df, float_col = [], str_col = []):

    if len(float_col) != 0:
        for col in float_col:
            # df.loc[:, col] = df[col].astype(str, errors = 'ignore')
            df.loc[:, col] = df[col].astype(str, errors='ignore')
            df.loc[:, col] = df[col].str.replace(',', '').astype(float, errors = 'ignore')
            df.loc[:, col] = pd.to_numeric(df[col], errors = 'coerce')
    if len(str_col) != 0:
        for col in str_col:
            df.loc[:, col] = df[col].astype(str, errors = 'ignore')
    return df

def combine_two_frames(df1, df2, col, suf = ('_Geneva', '_MUFG')):

    if df1 is None and df2 is None:
        return None

    if df1 is None:
        df1 = pd.DataFrame(columns = df2.columns)
        return combine_two_frames(df1, df2, col, suf)

    if df2 is None:
        df2 = pd.DataFrame(columns = df1.columns)
        return combine_two_frames(df1, df2, col, suf)

    df1 = df1.sort_values(by = [col]).reset_index(inplace=False, drop=True)
    df2 = df2.sort_values(by = [col]).reset_index(inplace=False, drop=True)

    res = pd.merge(df1, df2, left_index = True, right_index = True, how = 'outer', suffixes = suf)

    return res

#only for PR
def combine_two_frames2(df1, df2, col, suf = ('_Geneva', '_MUFG')):

    if df1 is None and df2 is None:
        return None

    if df1 is None:
        df1 = pd.DataFrame(columns = df2.columns)
        return combine_two_frames(df1, df2, col, suf)

    if df2 is None:
        df2 = pd.DataFrame(columns = df1.columns)
        return combine_two_frames(df1, df2, col, suf)

    df1 = df1.sort_values(by = col, ascending = False).reset_index(inplace=False, drop=True)
    df2 = df2.sort_values(by = col, ascending = False).reset_index(inplace=False, drop=True)

    res = pd.merge(df1, df2, left_index = True, right_index = True, how = 'outer', suffixes = suf)

    return res

#endregion

#region identical

def if_ident(df1, df2, idx1, idx2, identcols = []):

    if len(identcols) == 0:
        return True

    for i in identcols:

        if isinstance(df1.iloc[idx1][i], str) and isinstance(df2.iloc[idx2][i], str):

            if df1.iloc[idx1][i].upper() != df2.iloc[idx2][i].upper():
                return False

        elif isinstance(df1.iloc[idx1][i], numbers.Number) and isinstance(df2.iloc[idx2][i], numbers.Number):

            if df1.iloc[idx1][i] != df2.iloc[idx2][i]:
                return False

        else:
            print("errrrrrrrrrrrrrrrrrrrr")
            return False

    return True

#end region

class MyError(Exception):

    def __init__(self, value = None):
        self.value = value

    def __str__(self):
        return (repr(self.value))

class fuzzMatchTable():

    def __init__(self, df1 = pd.DataFrame(), df2 = pd.DataFrame()):
        self.idx_left = 'left idx'
        self.idx_right = 'right idx'
        self.df1 = df1.copy(deep = True)
        self.df2 = df2.copy(deep = True)

        self.left_all = 'left all'
        self.right_all = 'right all'
        self.leftprex = '_left'
        self.rightprex = '_right'

        self.df1_Matched = None
        self.df1_notMatched = None
        self.df2_notMatched = None

        self.left_idf = None
        self.right_idf = None

        if not df1.empty:
            self.df1.loc[:, 'left idx'] = df1.index.get_values()

        if not df2.empty:
            self.df2.loc[:, 'right idx'] = df2.index.get_values()

    def set_identifier(self, idf1, idf2):
        self.left_idf = idf1
        self.right_idf = idf2

    def set_left_identifier(self, idf1):
        self.left_idf = idf1

    def set_right_identifier(self, idf2):
        self.right_idf = idf2

    def match_name_dct(self, name, list_names, dct={}):

        if not isinstance(name, str):
            return '', -1, -1

        max_score = -1
        max_name = ""
        max_idx = -1

        for i in range(len(list_names)):

            if i in dct:
                continue

            name2 = list_names[i]

            # Finding fuzzy match score
            score = fuzz.ratio(name, name2)

            # Checking if we are above our threshold and have a better score
            if score > max_score:
                max_name = name2
                max_score = score
                max_idx = i

        return (max_name, max_score, max_idx)

    def match_name_dct_new(self, j, df1_ident, df2_ident, df1_idx, df2_idx, dct = {}, ifident = [], ver_cols = [], thred_cols = [], fuzzthred =100):

        name = df1_ident[j]

        if not isinstance(name, str):
            return '', -1, -1

        max_score = -1
        max_name = ""
        max_idx = -1

        for i in range(len(df2_ident)):

            if i in dct:
                continue

            if not if_ident(self.df1, self.df2, df1_idx[j], df2_idx[i], ifident):
                continue

            if not self.if_not_filter(self.df1, self.df2, df1_idx[j], df2_idx[i], ver_cols, thred_cols, fuzzthred):
                continue

            name2 = df2_ident[i]

            # Finding fuzzy match score
            score = fuzz.ratio(name, name2)

            # Checking if we are above our threshold and have a better score
            if score > max_score:
                max_name = name2
                max_score = score
                max_idx = i

        return (max_name, max_score, max_idx)


    def combine_strcol(self, df, mergecol=[]):

        if len(mergecol) == 0:
            return

        targcol = pd.DataFrame()

        try:
            for i in range(len(mergecol)):
                if i == 0:
                    targcol = df[mergecol[0]].replace(np.nan, '', regex = True).astype(str)
                else:
                    targcol = targcol + '--' + df[mergecol[i]].replace(np.nan, '', regex = True).astype(str)

        except KeyError as er:
            print(er.__repr__())
            exit(1)

        return targcol

    def comebine_strcol_all(self, mergecol = []):

        try:
            if len(mergecol) == 0:
                return

            self.df1.loc[:, self.left_all] = self.combine_strcol(self.df1, mergecol)
            self.df2.loc[:, self.right_all] = self.combine_strcol(self.df2, mergecol)

            self.df1_notMatched.loc[:, self.left_all] = self.combine_strcol(self.df1_notMatched, mergecol)
            self.df2_notMatched.loc[:, self.right_all] = self.combine_strcol(self.df2_notMatched, mergecol)
        except TypeError:
            pass

    def gen_merge_table(self, df1, df2, idx1, idx2, merge1, merge2, thred = 0, ifident = [], ver_cols = [], ver_thred = [], fuzzthred = 100):

        try:

            if merge1 not in df1.columns:
                raise MyError('df1 has no merge column')
            if merge2 not in df2.columns:
                raise MyError('df2 has no merge column')

            matched_dct = {}
            dct_list = []

            df1_ident = list(df1[merge1])
            df2_ident = list(df2[merge2])

            df1_idx = list(df1[idx1])
            df2_idx = list(df2[idx2])

            for i in range(len(df1_ident)):
                name = df1_ident[i]
                # match = self.match_name_dct(name, df2_ident, matched_dct)
                match = self.match_name_dct_new(i, df1_ident, df2_ident, df1_idx, df2_idx,
                                                matched_dct, ifident, ver_cols, ver_thred, fuzzthred)

                if match[1] == -1:
                    continue

                matched_dct[match[2]] = 1

                dict_ = {}
                dict_.update({"left idx": df1_idx[i]})
                dict_.update({"right idx": df2_idx[match[2]]})
                dict_.update({"left all": name})
                dict_.update({"right all": df2_ident[match[2]]})
                dict_.update({"score": match[1]})
                dct_list.append(dict_)

            merge_table = pd.DataFrame(dct_list)

            merge_table = merge_table[merge_table['score'] >= thred]

            return merge_table

        except KeyError as er:
            print(er)
            print('The comparison column must be initialized first')

        except MyError as er:
            print(er)

    def merge_with_table(self, df1, df2, idx1, idx2, mergetable):

        try:
            if df1.empty or df2.empty:
                raise MyError('table is empty')

            if mergetable is None:
                raise MyError('Merge table is None')

            if mergetable.empty:
                raise MyError('Merge table is empty')

            tempdf = pd.merge(df1, mergetable, how='inner', on=[idx1])
            tempdf = pd.merge(tempdf, df2, how='inner', on=[idx2], suffixes=(self.leftprex, self.rightprex))

            df1_un = df1[~df1.index.isin(tempdf[idx1])]
            df2_un = df2[~df2.index.isin(tempdf[idx2])]

            return tempdf, df1_un, df2_un

        except MyError as er:
            print(er)
            exit(1)

    def first_preload_match(self, flag = 0):

        # flag gives the option to set everything to default

        if (flag == 1) or ((self.df1_Matched is None) and (self.df2_notMatched is None) and (self.df1_notMatched is None)):
            self.df1_Matched = pd.DataFrame()
            self.df1_notMatched = self.df1.copy(deep=True)
            self.df2_notMatched = self.df2.copy(deep=True)

    def update_match(self, df_matched):

        self.df1_Matched = pd.concat([self.df1_Matched, df_matched], ignore_index=True, sort = False)

    def match_fuzz(self, thred = 0, vertcols = [], filterThred = [], fuzzThred = 100, ifident = []):

        # vertcols is used to filter the generated map

        try:

            self.first_preload_match()

            #region assert test
            assert self.df1_Matched.shape[0] + self.df1_notMatched.shape[0] == self.df1.shape[0]
            assert self.df1_Matched.shape[0] + self.df2_notMatched.shape[0] == self.df2.shape[0]
            #endregion

            #region not matched tables empty
            if self.df1_notMatched.empty or self.df2_notMatched.empty:
                print('match_fuzz test_not match table is empty now')
                return
            #endregion

            mergetable = self.gen_merge_table(self.df1_notMatched, self.df2_notMatched, self.idx_left,
                                              self.idx_right, self.left_all, self.right_all, thred, ifident, vertcols, filterThred, fuzzThred)

            #region mergetable empty
            if mergetable is None or mergetable.empty:
                print('merge table is empty now')
                return
            #endregion

            mergetable = mergetable[[self.idx_left, self.idx_right, 'score']]

            # renew with filterout
            # mergetable = self.filterout(self.df1_notMatched, self.df2_notMatched, mergetable,
            #                            vertcols, filterThred, fuzzThred)

            # region mergetable empty
            if mergetable is None or mergetable.empty:
                print('merge table is empty')
                return
            # endregion

            new_matches, self.df1_notMatched, self.df2_notMatched = self.merge_with_table(self.df1_notMatched, self.df2_notMatched, self.idx_left, self.idx_right, mergetable)

            # region assert test
            assert self.df1_Matched.shape[0] + new_matches.shape[0] + self.df1_notMatched.shape[0] == self.df1.shape[0]
            assert self.df1_Matched.shape[0] + new_matches.shape[0] + self.df2_notMatched.shape[0] == self.df2.shape[0]
            # endregion

            self.update_match(new_matches)

            # region assert test
            assert self.df1_Matched.shape[0] + self.df1_notMatched.shape[0] == self.df1.shape[0]
            assert self.df1_Matched.shape[0] + self.df2_notMatched.shape[0] == self.df2.shape[0]
            # endregion

        except AssertionError as er:
            print(er.__repr__())
            print("the matched rows plus unmatched rows should equal row number")
            exit(1)

        except KeyError as er:
            print(er.__repr__())
            exit(1)

    def super_two_stack_match(self, df1, df2, firstsort, thred=0, iden_cols = [], ver_cols = [], thred_cols = []):

        try:

            # region deal with lost idf
            if self.left_idf is None or self.right_idf is None:
                raise MyError('two stack match needs a second identifier, set it with set_identifier')
            # endregion

            df1 = df1[[firstsort, self.left_idf, self.idx_left]]
            df1 = df1.sort_values(by=[firstsort, self.left_idf], ascending=False)

            df2 = df2[[firstsort, self.right_idf, self.idx_right]]
            df2 = df2.sort_values(by=[firstsort, self.right_idf], ascending=False)

            stack1 = []

            er_stack1 = []
            er_stack2 = []

            df1_first = list(df1[firstsort])
            df1_idx = list(df1[self.idx_left])

            df2_first = list(df2[firstsort])
            df2_idx = list(df2[self.idx_right])

            track_dct = {}

            for i in range(len(df1_first)):

                ptr2 = 0

                while ptr2 <= len(df2_first):

                    if ptr2 in track_dct:
                        ptr2 += 1
                        continue

                    elif df1_first[i] == 0:
                        er_stack1.append(df1_idx[i])
                        break

                    # could be a problem
                    elif ptr2 < len(df2_first) and df1_first[i] < df2_first[ptr2] - abs(df2_first[ptr2]) * thred:
                        track_dct[ptr2] = 1
                        er_stack2.append(df2_idx[ptr2])

                    elif ptr2 == len(df2_first):
                        er_stack1.append(df1_idx[i])

                    elif (df1_first[i] >= df2_first[ptr2] - abs(df2_first[ptr2]) * thred) and (
                        df1_first[i] <= df2_first[ptr2] + abs(df2_first[ptr2]) * thred) and \
                        if_ident(self.df1, self.df2, df1_idx[i], df2_idx[ptr2], iden_cols) and \
                        self.if_not_filter(self.df1, self.df2, df1_idx[i], df2_idx[ptr2], ver_cols, thred_cols):

                        track_dct[ptr2] = 1
                        stack1.append((df1_idx[i], df2_idx[ptr2]))
                        break

                    else:
                        er_stack1.append(df1_idx[i])
                        break

                    ptr2 += 1


            for ii in range(len(df2_first)):
                if ii not in track_dct:
                    er_stack2.append(df2_idx[ii])


            assert len(stack1) + len(er_stack1) == df1.shape[0]
            assert len(stack1) + len(er_stack2) == df2.shape[0]

            if len(stack1) == 0:
                return None

            #region generate dataframe

            tempdf = pd.DataFrame([i[0] for i in stack1], columns=[self.idx_left])
            tempdf.loc[:, self.idx_right] = [i[1] for i in stack1]

            tempdf.loc[:, self.idx_left] = tempdf[self.idx_left].astype(int)
            tempdf.loc[:, self.idx_right] = tempdf[self.idx_right].astype(int)

            #endregion

            return tempdf

        except MyError as er:
            print(er)
            exit(1)

        except AssertionError as er:
            print(er.__repr__())
            exit(1)

    def match_stack(self, firstsort, thred = 0, vertcols = [], filterThred = [], fuzzThred = 100, idencols = []):

        try:
            # vertcols is used to filter the generated map

            self.first_preload_match()

            self.df1 = process_column_type(self.df1, vertcols)
            self.df2 = process_column_type(self.df2, vertcols)

            # region assert test
            assert self.df1_Matched.shape[0] + self.df1_notMatched.shape[0] == self.df1.shape[0]
            assert self.df1_Matched.shape[0] + self.df2_notMatched.shape[0] == self.df2.shape[0]
            # endregion

            # region not matched tables empty
            if self.df1_notMatched.empty or self.df2_notMatched.empty:
                print('match_stack notmatch table is empty now')
                return
            # endregion

            matchedmap = self.super_two_stack_match(self.df1_notMatched, self.df2_notMatched, firstsort, thred, idencols, vertcols, filterThred)

            if matchedmap is None or matchedmap.empty:
                print('stack match table is empty now')
                return

            # renew with filterout
            # matchedmap = self.filterout(self.df1_notMatched, self.df2_notMatched, matchedmap, vertcols, filterThred,
            #                                fuzzThred)

            # region mergetable empty
            #if matchedmap is None or matchedmap.empty:
            #    print('merge table is empty')
            #    return
            # endregion

            matched, self.df1_notMatched, self.df2_notMatched = self.merge_with_table(self.df1_notMatched, self.df2_notMatched, self.idx_left, self.idx_right, matchedmap)

            # region assert test
            assert self.df1_Matched.shape[0] + matched.shape[0] + self.df1_notMatched.shape[0] == self.df1.shape[0]
            assert self.df1_Matched.shape[0] + matched.shape[0] + self.df2_notMatched.shape[0] == self.df2.shape[0]
            # endregion

            self.update_match(matched)

            # region assert test
            assert self.df1_Matched.shape[0] + self.df1_notMatched.shape[0] == self.df1.shape[0]
            assert self.df1_Matched.shape[0] + self.df2_notMatched.shape[0] == self.df2.shape[0]
            # endregion

        except KeyError as er:
            print(er.__repr__())

    def getmyitem(self, df1, row, col):
        return df1.loc[[row], [col]].values[0][0]

    def filterout(self, df1, df2, map1, ver_cols, thredl, fuzzthred = 100):
        try:
            if len(ver_cols) == 0:
                return map1

            if len(ver_cols) != len(thredl):
                raise MyError("verify list must have the same length as thred list!")

            df1 = process_column_type(df1, ver_cols)
            df2 = process_column_type(df2, ver_cols)

            origin_len = map1.shape[0]

            er_stack1 = []
            er_stack2 = []

            for index, row in map1.iterrows():

                lftidx = row[self.idx_left]
                rgtidx = row[self.idx_right]

                if fuzz.ratio(self.getmyitem(df1, lftidx, self.left_all), self.getmyitem(df2, rgtidx, self.right_all)) >= fuzzthred:
                    continue

                for k in range(len(ver_cols)):
                    v = ver_cols[k]
                    thred = thredl[k]

                    try:
                        if abs(self.getmyitem(df2, rgtidx, v)) != 0:
                            if abs(self.getmyitem(df1, lftidx, v) - self.getmyitem(df2, rgtidx, v)) / abs(
                                    self.getmyitem(df2, rgtidx, v)) >= thred:
                                er_stack1.append(lftidx)
                                er_stack2.append(rgtidx)
                                map1.drop([index], inplace=True)
                                raise MyError()

                        elif abs(self.getmyitem(df1, lftidx, v)) != 0:
                            if abs(self.getmyitem(df1, lftidx, v) - self.getmyitem(df2, rgtidx, v)) / abs(
                                    self.getmyitem(df1, lftidx, v)) >= thred:
                                er_stack1.append(lftidx)
                                er_stack2.append(rgtidx)
                                map1.drop([index], inplace=True)
                                raise MyError()

                    except MyError:
                        break
                    except:
                        pass

            assert map1.shape[0] + len(er_stack1) == origin_len

            map1.reset_index(inplace = True, drop = True)

            return map1

        except MyError as er:
            print(er)
            exit(1)

    def fillscore(self):
        if self.df1_Matched is not None and not self.df1_Matched.empty:
            self.df1_Matched.loc[:, 'score'] = self.df1_Matched.apply(lambda row: fuzz.ratio(row[self.left_all], row[self.right_all]), axis=1)

    def caldiff(self, cols):

        if self.df1_Matched is None or self.df1_Matched.empty:
            return

        tempcols = []
        for l in cols:
            tempcols.append(l + self.leftprex)
            tempcols.append(l + self.rightprex)

        self.df1_Matched = process_column_type(self.df1_Matched, tempcols)

        for l in cols:
            self.df1_Matched.loc[:, l + '_diff'] = self.df1_Matched[l + self.leftprex] - self.df1_Matched[l + self.rightprex]

    def super_match1(self, fuzzcols, targcol, targthred, secondsort, vertcols = [], thred = [], DIVTIME = 0, idencols = []):

        try:

            self.first_preload_match()

            if len(vertcols) != len(thred):
                raise MyError('vertcols and thred should be of same length')

            self.df1 = process_column_type(self.df1, [targcol] + vertcols)
            self.df2 = process_column_type(self.df2, [targcol] + vertcols)

            self.df1_notMatched = process_column_type(self.df1_notMatched, [targcol] + vertcols)
            self.df2_notMatched = process_column_type(self.df2_notMatched, [targcol] + vertcols)

            self.comebine_strcol_all(fuzzcols)
            self.set_identifier(secondsort, secondsort)

            temp_thred = [0 for i in range(len(thred))]

            if DIVTIME == 0:
                self.match_stack(targcol, targthred, vertcols, temp_thred, idencols=idencols)
                return

            for i in range(DIVTIME):

                if len(vertcols) == 0:
                    self.match_stack(targcol, targthred, vertcols, temp_thred, idencols=idencols)
                else:
                    for j in range(len(vertcols)-1, -1, -1):
                        step = thred[j] / DIVTIME
                        temp_thred[j] += step

                        print(temp_thred)

                        self.match_stack(targcol, targthred, vertcols, temp_thred, idencols = idencols)

            self.fillscore()

        except MyError as er:
            print(er)
            exit(1)

        except TypeError as er:
            print(er.__repr__())
            pass

    def super_match2(self, fuzzcols, targcol, targthred, secondsort, vertcols=[], thred=[],
                     DIVTIME1=10, DIVTIME2 = 10, FUZZthred = 10, idencols = []):

        try:

            self.first_preload_match()

            if len(vertcols) != len(thred):
                raise MyError('vertcols and thred should be of same length')

            self.comebine_strcol_all(fuzzcols)

            step = targthred / DIVTIME1
            start = 0
            FUZZ_start = 99
            FUZZ_step = FUZZthred / DIVTIME1

            for i in range(DIVTIME1 + 1):

                print(start)

                self.match_fuzz(FUZZ_start, ifident = idencols)

                self.super_match1(fuzzcols, targcol, start, secondsort, vertcols, thred, DIVTIME2, idencols)

                start += step
                FUZZ_start -= FUZZ_step


            self.fillscore()

            self.caldiff([targcol] + vertcols)

        except MyError as er:
            print(er)
            exit(1)


    def super_match3(self, fuzzcols, targcol, targthred, secondsort, vertcols=[], thred=[],
                     DIVTIME1=10, DIVTIME2 = 10, FUZZthred = 10, idencols = []):

        try:

            self.first_preload_match()

            if len(vertcols) != len(thred):
                raise MyError('vertcols and thred should be of same length')

            self.comebine_strcol_all(fuzzcols)

            step = targthred / DIVTIME1
            start = 0

            for i in range(DIVTIME1 + 1):

                self.super_match1(fuzzcols, targcol, start, secondsort, vertcols, thred, DIVTIME2, idencols)
                start += step

            # self.match_stack(targcol, 0)

            self.fillscore()

            self.caldiff([targcol] + vertcols)

        except MyError as er:
            print(er)
            exit(1)

    def if_not_filter(self, df1, df2, idx1, idx2, ver_cols=[], thredl=[], fuzzthred=100):

        try:
            if len(ver_cols) == 0:
                return True

            if len(ver_cols) != len(thredl):
                raise MyError("verify list must have the same length as thred list!")

            # should be repalced
            # df1 = process_column_type(df1, ver_cols)
            # df2 = process_column_type(df2, ver_cols)

            if fuzz.ratio(df1.iloc[idx1][self.left_all], df2.iloc[idx2][self.right_all]) >= fuzzthred:
                return True

            for i in range(len(ver_cols)):

                v = ver_cols[i]
                thred = thredl[i]

                if abs(df1.iloc[idx1][v]) != 0:
                    if abs(df1.iloc[idx1][v] - df2.iloc[idx2][v]) / abs(df1.iloc[idx1][v]) > thred:
                        return False

                elif abs(df2.iloc[idx2][v]) != 0:
                    if abs(df1.iloc[idx1][v] - df2.iloc[idx2][v]) / abs(df2.iloc[idx2][v]) > thred:
                        return False

            return True

        except MyError as er:
            print(er)
            exit(1)


class TestStrMatch(unittest.TestCase):

    testClass = fuzzMatchTable()

    def assertDF(self, df1, df2):
        self.assertDictEqual(df1.to_dict(), df2.to_dict())

    def testif_ident(self):
        df1_cons = pd.DataFrame({'a': ['we have', 'you have'], 'b': ['a dream', 'two dreams'], 'c': ['a dream', 'two dreams']})
        df2_cons = pd.DataFrame({'a': ['we have', 'you hav'], 'b': ['a dream', 'two dreams'], 'c': ['two dream', 'two dreams']})

        self.assertTrue(if_ident(df1_cons, df2_cons, 0, 0, ['a']))
        self.assertTrue(not if_ident(df1_cons, df2_cons, 1, 1, ['a']))

    def test_constructor(self):
        df1_cons = pd.DataFrame({'a': ['we have', 'you have'], 'b': ['a dream', 'two dreams']})
        df2_cons = pd.DataFrame({'a': ['we hav', 'you hav'], 'b': ['a dream', 'two dreams']})

        testClass_constructor = fuzzMatchTable(df1_cons, df2_cons)

        self.assertDictEqual(df1_cons[['a', 'b']].to_dict(), testClass_constructor.df1[['a', 'b']].to_dict())
        self.assertTrue(testClass_constructor.df1['left idx'] is not None)

    def sudodate(self):
        return datetime.strptime('10/31/2018', '%m/%d/%Y')

    def test_match_dct(self):
        self.assertEqual(self.testClass.match_name_dct('jack', ['jack', 'jack', 'jack'], {}), ('jack', 100, 0))
        self.assertEqual(self.testClass.match_name_dct('jack pak', ['jac block', 'acy pa', 'jack park'], {}), ('jack park', 94, 2))
        self.assertEqual(self.testClass.match_name_dct('jack pak', ['jac block', 'acy pa', 'jack park'], {2:1}), ('acy pa', 71, 1))
        self.assertEqual(self.testClass.match_name_dct('jack pak', [], {}), ('', -1, -1))

    def test_combine_strcol(self):
        df1 = pd.DataFrame({'a':['we have', 'you have'], 'b':['a dream', 'two dreams']})
        df1_test = pd.DataFrame({'a': ['we have', 'you have'], 'b': ['a dream', 'two dreams'], 'c':['we have--a dream', 'you have--two dreams']})

        df1['c'] = self.testClass.combine_strcol(df1, ['a', 'b'])
        self.assertDictEqual(df1.to_dict(), df1_test.to_dict())

        df2 = self.testClass.combine_strcol(df1, ['a'])
        df2_test = df1['a']
        self.assertDictEqual(df2.to_dict(), df2_test.to_dict())

        df3 = pd.DataFrame({'a': [1, 2], 'b': [1, 2]})
        df3['d'] = pd.DataFrame({'d': ['1--1', '2--2']})
        df3['c'] = self.testClass.combine_strcol(df3, ['a', 'b'])
        self.assertDictEqual(df3['c'].to_dict(), df3['d'].to_dict())

        df4 = pd.DataFrame({'a': [np.nan, np.nan], 'b': [np.nan, 2]})
        df4['c'] = self.testClass.combine_strcol(df4, ['a', 'b'])
        df4['d'] = pd.DataFrame({'d': ['--', '--2.0']})
        # all -- needs to be filtered out
        self.assertDictEqual(df4['c'].to_dict(), df4['d'].to_dict())

    def test_if_not_filter(self):
        df1_cons = pd.DataFrame({'name': ['jack', 'paul'], 'income': [21.5, 320]})
        df2_cons = pd.DataFrame({'name': ['jac', 'pau'], 'income': [300, 21.5]})

        tc = fuzzMatchTable(df1_cons, df2_cons)
        tc.comebine_strcol_all(['name'])

        df1_cons = tc.df1
        df2_cons = tc.df2

        temp = tc.if_not_filter(df1_cons, df2_cons, 0, 1, ['income'], [0])
        temp2 = tc.if_not_filter(df1_cons, df2_cons, 1, 0, ['income'], [0.06])
        temp3 = tc.if_not_filter(df1_cons, df2_cons, 1, 0, ['income'], [0.07])

        self.assertTrue(temp)
        self.assertTrue(not temp2)
        self.assertTrue(temp3)

    def test_combine_strcol_all(self):
        df1_cons = pd.DataFrame({'a': ['we have', 'you have'], 'b': ['a dream', 'two dreams']})
        df2_cons = pd.DataFrame({'a': ['we hav', 'you hav'], 'b': ['a dream', 'two dreams']})
        df1_test = pd.DataFrame({'a': ['we have', 'you have'], 'b': ['a dream', 'two dreams'], 'c': ['we have--a dream', 'you have--two dreams']})
        df2_test = pd.DataFrame({'a': ['we hav', 'you hav'], 'b': ['a dream', 'two dreams'], 'c': ['we hav--a dream', 'you hav--two dreams']})

        testClass_constructor = fuzzMatchTable(df1_cons, df2_cons)

        testClass_constructor.comebine_strcol_all(['a', 'b'])

        self.assertDictEqual(testClass_constructor.df1['left all'].to_dict(), df1_test['c'].to_dict())
        self.assertDictEqual(testClass_constructor.df2['right all'].to_dict(), df2_test['c'].to_dict())

    def test_gen_merge_table(self):

        # region test mergetable
        df1_cons = pd.DataFrame({'a': ['we have', 'you have'], 'b': ['a dream', 'two dreams']})
        df2_cons = pd.DataFrame({'a': ['we hav', 'you hav'], 'b': ['an apple', 'two dreams']})
        df1_test = pd.DataFrame({'a': ['we have', 'you have'], 'b': ['a dream', 'two dreams'],
                                 'c': ['we have--a dream', 'you have--two dreams']})

        df1_cons['c'] = self.testClass.combine_strcol(df1_cons, ['a', 'b'])
        df2_cons['c'] = self.testClass.combine_strcol(df2_cons, ['a', 'b'])

        df1_cons['left idx'] = df1_cons.index.get_values()
        df2_cons['right idx'] = df2_cons.index.get_values()

        mergetable = self.testClass.gen_merge_table(df1_cons, df2_cons, 'left idx', 'right idx', 'c', 'c')[['left idx', 'right idx']]

        res_table = pd.DataFrame({'left idx': [0, 1], 'right idx': [0, 1]})

        self.assertEqual(res_table.to_dict(), mergetable.to_dict())

        mergetable2 = self.testClass.gen_merge_table(df1_cons, df2_cons, 'left idx', 'right idx', 'c', 'c', 100)[
            ['left idx', 'right idx']]

        self.assertTrue(mergetable2.empty)

        #endregion

        #region test merget_with_table
        mergetable3 = self.testClass.gen_merge_table(df1_cons, df2_cons, 'left idx', 'right idx', 'c', 'c', 75)

        matched_temp = self.testClass.merge_with_table(df1_cons, df2_cons, 'left idx', 'right idx', mergetable3)[0][['left idx', 'right idx']]

        res_table2 = pd.DataFrame({'left idx': [1], 'right idx': [1]})

        self.assertDictEqual(res_table2.to_dict(), matched_temp.to_dict())
        #endregion

    def test_gen_merge_table2(self):

        df1_cons = pd.DataFrame({'a': ['we have', 'you have'], 'b': ['a dream', 'two dreams'], 'd':['word', 'word']})
        df2_cons = pd.DataFrame({'a': ['we hav', 'you hav'], 'b': ['an apple', 'two dreams'], 'd':['wrd', 'word']})

        df1_cons['left idx'] = df1_cons.index.get_values()
        df2_cons['right idx'] = df2_cons.index.get_values()

        testClass2 = fuzzMatchTable(df1_cons, df2_cons)
        testClass2.comebine_strcol_all(['a', 'b'])

        df1_cons = testClass2.df1
        df2_cons = testClass2.df2

        mergetable = testClass2.gen_merge_table(df1_cons, df2_cons, 'left idx', 'right idx', 'left all', 'right all', ifident= ['d'])[
            ['left idx', 'right idx']]


        res_table = pd.DataFrame({'left idx': [0], 'right idx': [1]})

        self.assertEqual(res_table.to_dict(), mergetable.to_dict())


    def test_match_fuzz(self):

        df1_cons = pd.DataFrame({'a': ['we have', 'you have'], 'b': ['a dream', 'two dreams']})
        df2_cons = pd.DataFrame({'a': ['we hav', 'you hav'], 'b': ['an apple', 'two dreams']})

        testClass_constructor = fuzzMatchTable(df1_cons, df2_cons)
        testClass_constructor.comebine_strcol_all(['a', 'b'])

        testClass_constructor.match_fuzz()

        res_table = pd.DataFrame({'left idx': [0, 1], 'right idx': [0, 1]})

        self.assertDictEqual(testClass_constructor.df1_Matched[['left idx', 'right idx']].to_dict(), res_table.to_dict())

        testClass_constructor.match_fuzz(75)

        self.assertDictEqual(testClass_constructor.df1_Matched[['left idx', 'right idx']].to_dict(), res_table.to_dict())

        testClass_constructor.first_preload_match(1)

        testClass_constructor.match_fuzz(75)

        res_table = pd.DataFrame({'left idx': [1], 'right idx': [1]})

        self.assertDictEqual(testClass_constructor.df1_Matched[['left idx', 'right idx']].to_dict(), res_table.to_dict())

        testClass_constructor.match_fuzz(0)

        res_table = pd.DataFrame({'left idx': [1, 0], 'right idx': [1, 0]})

        self.assertDictEqual(testClass_constructor.df1_Matched[['left idx', 'right idx']].to_dict(), res_table.to_dict())

    def test_match_fuzz2(self):

        df1_cons = pd.DataFrame({'a': ['we have', 'you have'], 'b': ['a dream', 'two dreams'], 'c': ['oe', 'one']})
        df2_cons = pd.DataFrame({'a': ['we hav', 'you hav'], 'b': ['an apple', 'two dreams'], 'c': ['one', 'one']})

        testClass_constructor = fuzzMatchTable(df1_cons, df2_cons)
        testClass_constructor.comebine_strcol_all(['a', 'b'])

        testClass_constructor.match_fuzz(ifident=['c'])

        res_table = pd.DataFrame({'left idx': [1], 'right idx': [1]})

        self.assertDictEqual(testClass_constructor.df1_Matched[['left idx', 'right idx']].to_dict(), res_table.to_dict())


    def test_super_two_stack_match(self):

        #region first test
        df1_cons = pd.DataFrame({'name': ['jack', 'paul'], 'income': [21.5, 300]})
        df2_cons = pd.DataFrame({'name': ['jac', 'pau'], 'income': [300, 21.5]})

        tc = fuzzMatchTable(df1_cons, df2_cons)
        tc.set_identifier('name', 'name')

        test_res1 = pd.DataFrame({'left idx': [1, 0], 'right idx': [0, 1]})
        res1 = tc.super_two_stack_match(tc.df1, tc.df2, 'income')

        self.assertDictEqual(test_res1.to_dict(), res1.to_dict())
        #endregion

        #region second test
        df2_cons = pd.DataFrame({'name': ['jac', 'pau'], 'income': [300, 100]})
        tc = fuzzMatchTable(df1_cons, df2_cons)
        tc.set_identifier('name', 'name')

        test_res2 = res_table = pd.DataFrame({'left idx': [1], 'right idx': [0]})
        res2 = tc.super_two_stack_match(tc.df1, tc.df2, 'income', 0.1)

        self.assertDictEqual(test_res2.to_dict(), res2.to_dict())
        #endregion second test

        # region third test
        df2_cons = pd.DataFrame({'name': ['jac', 'pau'], 'income': [300, 21]})
        tc = fuzzMatchTable(df1_cons, df2_cons)
        tc.set_identifier('name', 'name')

        test_res2 = pd.DataFrame({'left idx': [1, 0], 'right idx': [0, 1]})
        res2 = tc.super_two_stack_match(tc.df1, tc.df2, 'income', 0.05)

        self.assertDictEqual(test_res2.to_dict(), res2.to_dict())
        # endregion second test

    def test_super_two_stack_match2(self):
        # region first test
        df1_cons = pd.DataFrame({'name': ['jack', 'paul'], 'income': [21.5, 300], 'job':["we", "we"]})
        df2_cons = pd.DataFrame({'name': ['pau', 'jack'], 'income': [300, 21.5], 'job':["we", "we"]})

        tc = fuzzMatchTable(df1_cons, df2_cons)
        tc.set_identifier('name', 'name')

        res1 = tc.super_two_stack_match(tc.df1, tc.df2, 'income', iden_cols = ['name', 'job'])

        test_res1 = pd.DataFrame({'left idx': [0], 'right idx': [1]})

        self.assertDictEqual(test_res1.to_dict(), res1.to_dict())

    def test_super_two_stack_match3(self):
        # region first test
        df1_cons = pd.DataFrame({'name': ['jack_1', 'paul'], 'income': [21.5, 320], 'nincome':[320, 300],'job':["we", "we"]})
        df2_cons = pd.DataFrame({'name': ['pau', 'jack'], 'income': [300, 21.5], 'nincome':[320, 300], 'job':["we", "we"]})

        tc = fuzzMatchTable(df1_cons, df2_cons)
        tc.set_identifier('name', 'name')
        tc.comebine_strcol_all(['name'])

        res1 = tc.super_two_stack_match(tc.df1, tc.df2, 'income', ver_cols = ['nincome'], thred_cols = [0.07])

        test_res1 = pd.DataFrame({'left idx': [0], 'right idx': [1]})

        self.assertDictEqual(test_res1.to_dict(), res1.to_dict())


    def test_match_stack(self):

        #region test1
        df1_cons = pd.DataFrame({'name': ['jack', 'paul'], 'income': [21.5, 300]})
        df2_cons = pd.DataFrame({'name': ['jac', 'pau'], 'income': [300, 21.5]})

        tc = fuzzMatchTable(df1_cons, df2_cons)
        tc.set_identifier('name', 'name')

        tc.match_stack('income', 0)

        test_res1 = pd.DataFrame({'left idx': [0, 1], 'right idx': [1, 0]})

        self.assertDF(test_res1, tc.df1_Matched[['left idx', 'right idx']])
        #endregion

    def test_match_stack2(self):

        # region test1

        df1_cons = pd.DataFrame({'name': ['jack', 'Perk', 'luke'], 'income': [1000, 5000, 400], 'speed':[50, 40, 20]})
        df2_cons = pd.DataFrame({'name': ['jac', 'luke', 'perk1'], 'income': [990, 5000, 5000], 'speed':[50, 40, 20]})

        tc = fuzzMatchTable(df1_cons, df2_cons)
        tc.set_identifier('name', 'name')

        tc.comebine_strcol_all(['name'])
        tc.match_fuzz(99)

        matched = tc.df1_Matched[['left idx', 'right idx']]
        matched_test1 = pd.DataFrame({'left idx':[2], 'right idx':[1]})
        self.assertDF(matched, matched_test1)

        tc.match_stack('income')
        matched2 = tc.df1_Matched[['left idx', 'right idx']]
        matched_test2 = pd.DataFrame({'left idx': [2, 1], 'right idx': [1, 2]})
        self.assertDF(matched2, matched_test2)

        tc.match_stack('speed')
        matched3 = tc.df1_Matched[['left idx', 'right idx']]
        matched_test3 = pd.DataFrame({'left idx': [2, 1, 0], 'right idx': [1, 2, 0]})
        self.assertDF(matched3, matched_test3)

        # end region

    def test_filterout(self):

        df1_cons = pd.DataFrame({'name': ['jack', 'Perk', 'luke'], 'income': [1000, 5000, 400], 'speed': [50, 40, 20]})
        df2_cons = pd.DataFrame({'name': ['jac', 'luke', 'perk1'], 'income': [990, 5000, 5000], 'speed': [50, 40, 20]})

        tc = fuzzMatchTable(df1_cons, df2_cons)
        tc.set_identifier('name', 'name')
        tc.comebine_strcol_all(['name'])

        matched_map = pd.DataFrame({'left idx': [2, 1, 0], 'right idx': [1, 2, 0]})

        map_test = self.testClass.filterout(tc.df1, tc.df2, matched_map, ['income'], [0.009], 101)
        test1 = pd.DataFrame({'left idx': [1], 'right idx': [2]})

        self.assertDF(test1, map_test)

        matched_map = pd.DataFrame({'left idx': [2, 1, 0], 'right idx': [1, 2, 0]})

        map_test = self.testClass.filterout(tc.df1, tc.df2, matched_map, ['income'], [0.1], 101)
        test1 = pd.DataFrame({'left idx': [1, 0], 'right idx': [2, 0]})

        self.assertDF(test1, map_test)

    def test_match_fuzz_with_vert(self):

        df1_cons = pd.DataFrame({'name': ['jack', 'Perk', 'luke'], 'income': [1000, 5000, 400], 'speed': [50, 40, 20]})
        df2_cons = pd.DataFrame({'name': ['jac', 'luke', 'perk1'], 'income': [990, 5000, 5000], 'speed': [50, 40, 20]})

        tc = fuzzMatchTable(df1_cons, df2_cons)
        tc.set_identifier('name', 'name')
        tc.comebine_strcol_all(['name'])

        tc.match_fuzz(100, ['income'], [0], 101)
        test1 = tc.df1_Matched

        self.assertTrue(test1.empty)

        tc.match_fuzz(50, ['income'], [0.1], 101)
        test1 = tc.df1_Matched[['left idx', 'right idx']]

        test2 = pd.DataFrame({'left idx': [0, 1], 'right idx': [0, 2]})

        self.assertDF(test1, test2)

    def test_match_stack_with_vert(self):

        df1_cons = pd.DataFrame({'name': ['jack', 'Perk', 'luke'], 'income': [1000, 5000, 400], 'speed': [50, 40, 20]})
        df2_cons = pd.DataFrame({'name': ['jac', 'luke', 'perk1'], 'income': [990, 5000, 5000], 'speed': [50, 40, 20]})

        tc = fuzzMatchTable(df1_cons, df2_cons)
        tc.set_identifier('name', 'name')
        tc.comebine_strcol_all(['name'])

        tc.match_stack('income', 0.1, ['income', 'speed'], [0.1, 0.1])

        test1 = tc.df1_Matched[['left idx', 'right idx']]

        test2 = pd.DataFrame({'left idx': [0], 'right idx': [0]})

        self.assertDF(test1, test2)


if __name__ == "__main__":

    unittest.main()


